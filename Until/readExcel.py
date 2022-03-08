#coding=utf-8
import os
import getpathInfo #自己定义的内部类，该类返回项目的绝对路径
#调用读Excel的第三方库xlrd
from xlrd import open_workbook
import openpyxl
import xlwt
#拿到该项目所在的绝对路径
path = getpathInfo.get_Path()


class readExcel():
    def get_xls(self, xls_name, sheet_name):
        """获取所有数据 open_workbook"""
        """open_workbook下标从0计算，而openpyxl下标是从1开始"""
        cls = []
        #获取用例文件路径
        xlsPath = os.path.join(path, 'testFile',  xls_name)
        file = open_workbook(xlsPath)
        sheet = file.sheet_by_name(sheet_name)
        nrows = sheet.nrows
        for i in range(nrows):
            if sheet.row_values(i)[1] != u'case_name':
                cls.append(sheet.row_values(i))
        return cls

    def load_excel(self, excelfile=None):
        """
        加载Excel文件  openpyxl
        :return:
        """
        if excelfile == None:
            excelfile = "userCase.xlsx"
        excelPath = "/testFile/" + excelfile
        excelFiledata = openpyxl.load_workbook(path + excelPath)
        return excelFiledata


    def get_sheet_data(self, excelfile=None, sheetName=None):
        """
        获取指定sheet的内容
        :param index:
        :return:
        """
        if sheetName == None:
            sheetName = "login"
        data = self.load_excel(excelfile).get_sheet_by_name(sheetName)
        return data

    def get_cell_value(self, row, cols, excelfile, sheetname):
        """
        获取某一个单元格内容
        :param row:
        :param cols:
        :return:
        """
        data = self.get_sheet_data(excelfile, sheetname).cell(row=row, column=cols+1).value
        return data

    def get_rows(self,excelfile=None, sheetName=None):
        """
        获取行数
        :return:
        """
        row = self.get_sheet_data(excelfile, sheetName).max_row
        return row

    def get_rows_value(self, row):
        """
        获取某一行的内容
        :param row:
        :return:
        """
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value, excelfile=None):
        """
        写入数据
        :param row:
        :param cols:
        :param value:
        :return:
        """
        if excelfile == None:
            excelfile = "userCase.xlsx"
        excelPath = "/testFile/" + excelfile
        wb = self.load_excel(excelfile)
        wr = wb.active
        wr.cell(row, cols+1, value)
        wb.save(path + excelPath)


    def get_columns_value(self, key=None, excelfile=None, sheetName=None):
        """
        获取某一列
        :param key:
        :return:
        """
        columns_list = []
        if key == None:
            key = 'A'
        columns_list_data = self.get_sheet_data(excelfile, sheetName)[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id, excelfile=None, sheetName=None):
        """获取行号"""
        num = 1
        cols_data = self.get_columns_value(excelfile=excelfile, sheetName=sheetName)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1

    def get_excel_data(self):
        """获取Excel里所有的数据"""
        data_list = []
        for i in range(self.get_rows() - 1):
            data_list.append(self.get_rows_value(i + 2))
        return data_list

excel_data = readExcel()

if __name__ == '__main__':
    data = readExcel().get_xls('userCase.xlsx', 'login')
    # print(readExcel().excel_write_data(6,2,"sadas"))
    # print(readExcel().get_xls('userCase.xlsx', 'login'))
    # print(readExcel().get_xls('userCase.xlsx', 'login')[0][1])
    # print(readExcel().get_xls('userCase.xlsx', 'login')[1][2])
    # print(type(data))
    # data = readExcel().load_excel().sheetnames
    # print(readExcel().load_excel()["login"])
    # print(data[0])

    # ##清空 resData两列数据
    # rowdums = excel_data.get_rows("cbk_user.xlsx", "commodityPurchase")
    # for i in range(2, rowdums+1):
    #     excel_data.excel_write_data(i, 14, "", "cbk_user.xlsx")
    #     excel_data.excel_write_data(i, 15, "", "cbk_user.xlsx")




    #使用xlrd
    # def get_excelFile(self, excel_name=None):
    #     """获取excel文件"""
    #     if excel_name == None:
    #         excel_name = 'userCase.xlsx'
    #     excelpath = os.path.join(path, 'testFile', excel_name)
    #     file = open_workbook(excelpath)
    #     return file
    #
    # def get_excelSheet(self, sheet_index=0):
    #     """获取对应sheet表"""
    #     file = self.get_excelFile()
    #     sheet = file.sheet_by_index(sheet_index)
    #     return sheet
    #
    # def get_rows(self):
    #     """获取总行数"""
    #     rows = self.get_excelSheet().nrows
    #     return rows
    #
    # def get_row_data(self, n):
    #     """获取一行数据"""
    #     sheet_data = self.get_excelSheet()
    #     row_data = sheet_data.row_values(n)
    #     return row_data
    #
    # def get_cell(self, row, cols):
    #     """获取单元格内容"""
    #     # row_data = self.get_excelSheet().row_values(row)[clos]
    #     cell_data = self.get_excelSheet().cell(row, cols).value
    #     return cell_data
    #
    # def get_cols(self, cols):
    #     """获取列数"""
    #     sheet_data = self.get_excelSheet()
    #     cols_data = sheet_data.col_slice(cols)
    #     return cols_data



